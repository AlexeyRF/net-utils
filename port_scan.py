import socket
import concurrent.futures
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
import os
from datetime import datetime
import ipaddress
def get_service_name(port):
    try:
        return socket.getservbyport(port)
    except:
        return "Неизвестно"
def scan_port(ip, port, timeout=1):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((ip, port))
        sock.close()
        if result == 0:
            service = get_service_name(port)
            return port, "Открыт", service
        else:
            return port, "Закрыт", ""
    except socket.timeout:
        return port, "Таймаут", ""
    except socket.error:
        return port, "Ошибка", ""
    except Exception as e:
        return port, f"Ошибка: {str(e)}", ""
def scan_ports(ip, start_port=1, end_port=65535, max_threads=500, timeout=1):
    open_ports = []
    print(f"\nНачинаем сканирование портов {ip}:{start_port}-{end_port}")
    print(f"Таймаут: {timeout} сек, Потоков: {max_threads}")
    print("-" * 40)
    total_ports = end_port - start_port + 1
    scanned_ports = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_threads) as executor:
        futures = {
            executor.submit(scan_port, ip, port, timeout): port
            for port in range(start_port, end_port + 1)
        }
        for future in concurrent.futures.as_completed(futures):
            scanned_ports += 1
            port, status, service = future.result()
            progress = (scanned_ports / total_ports) * 100
            if scanned_ports % 1000 == 0 or scanned_ports == total_ports:
                print(f"Прогресс: {scanned_ports}/{total_ports} ({progress:.1f}%) - "
                      f"Найдено открытых: {len(open_ports)}", end='\r')
            if status == "Открыт":
                open_ports.append((port, service))
                print(f"  Порт {port}: ОТКРЫТ ({service})")
    print(f"\nСканирование завершено!")
    print(f"Найдено открытых портов: {len(open_ports)} из {total_ports}")
    print(f"Диапазон: {start_port}-{end_port}")
    return open_ports
def get_host_info(ip):
    info = {
        "IP": ip,
        "Сканировано": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Имя хоста": "",
        "Статус": ""
    }
    try:
        hostname, _, _ = socket.gethostbyaddr(ip)
        info["Имя хоста"] = hostname
    except:
        info["Имя хоста"] = "Не удалось определить"
    try:
        socket.create_connection((ip, 80), timeout=2)
        info["Статус"] = "Хост доступен"
    except:
        info["Статус"] = "Хост не отвечает"
    return info
def save_to_excel(ip, open_ports, host_info, port_range, filename=None):
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"port_scan_{ip}_{timestamp}.xlsx"
    wb = Workbook()
    if open_ports:
        ws_ports = wb.active
        ws_ports.title = "Открытые порты"
        headers = ["Порт", "Служба", "Описание", "Риск", "Рекомендации"]
        ws_ports.append(headers)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_ports[1]:
            cell.fill = header_fill
            cell.font = header_font
        common_ports = {
            21: ("FTP", "Высокий", "Используйте SFTP/FTPS"),
            22: ("SSH", "Средний", "Используйте ключи вместо паролей"),
            23: ("Telnet", "Высокий", "Замените на SSH"),
            25: ("SMTP", "Средний", "Настройте аутентификацию"),
            80: ("HTTP", "Низкий", "Перейдите на HTTPS"),
            110: ("POP3", "Средний", "Используйте SSL/TLS"),
            139: ("NetBIOS", "Высокий", "Закройте если не используется"),
            443: ("HTTPS", "Низкий", "Проверьте сертификат"),
            445: ("SMB", "Высокий", "Обновите до последней версии"),
            3389: ("RDP", "Высокий", "Используйте VPN"),
            5900: ("VNC", "Высокий", "Используйте SSH туннель"),
        }
        for port, service in open_ports:
            port_info = common_ports.get(port, ("Неизвестно", "Не определен", "Требуется анализ"))
            ws_ports.append([port, service, port_info[0], port_info[1], port_info[2]])
    ws_info = wb.create_sheet(title="Информация")
    ws_info.append(["Результаты сканирования портов"])
    ws_info.append([])
    for key, value in host_info.items():
        ws_info.append([key, value])
    ws_info.append(["Диапазон портов", f"{port_range[0]}-{port_range[1]}"])
    ws_info.append(["Всего проверено", f"{port_range[1] - port_range[0] + 1}"])
    ws_info.append(["Найдено открытых", f"{len(open_ports)}"])
    if open_ports:
        percentage = (len(open_ports) / (port_range[1] - port_range[0] + 1)) * 100
        ws_info.append(["Процент открытых", f"{percentage:.2f}%"])
    ws_info.append([])
    ws_stats = wb.create_sheet(title="Статистика")
    categories = {
        "Системные (0-1023)": 0,
        "Пользовательские (1024-49151)": 0,
        "Динамические (49152-65535)": 0
    }
    for port, _ in open_ports:
        if port <= 1023:
            categories["Системные (0-1023)"] += 1
        elif port <= 49151:
            categories["Пользовательские (1024-49151)"] += 1
        else:
            categories["Динамические (49152-65535)"] += 1
    ws_stats.append(["Категория портов", "Количество", "Процент"])
    total_ports_scanned = port_range[1] - port_range[0] + 1
    for category, count in categories.items():
        if total_ports_scanned > 0:
            percentage = (count / total_ports_scanned) * 100
            ws_stats.append([category, count, f"{percentage:.2f}%"])
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(filename)
    return filename
def main():
    print("=" * 40)
    print("СКАНЕР ПОРТОВ v1.0")
    print("=" * 40)
    while True:
        ip = input("\nВведите IP-адрес для сканирования: ").strip()
        try:
            ipaddress.ip_address(ip)
            break
        except ValueError:
            print("Ошибка: Неверный формат IP-адреса. Попробуйте снова.")
    print("\nНастройки сканирования:")
    print("-" * 40)
    print("\nДиапазон портов для сканирования:")
    print("1. Все порты (1-65535)")
    print("2. Только известные порты (1-1023)")
    print("3. Пользовательский диапазон")
    choice = input("Выберите вариант (1-3): ").strip()
    if choice == '1':
        start_port, end_port = 1, 65535
    elif choice == '2':
        start_port, end_port = 1, 1023
    else:
        try:
            start_port = int(input("Начальный порт (1-65535): ").strip())
            end_port = int(input("Конечный порт (1-65535): ").strip())
            if not (1 <= start_port <= 65535 and 1 <= end_port <= 65535):
                print("Порты должны быть в диапазоне 1-65535. Использую значения по умолчанию.")
                start_port, end_port = 1, 65535
            if start_port > end_port:
                start_port, end_port = end_port, start_port
        except:
            print("Ошибка ввода. Использую все порты (1-65535).")
            start_port, end_port = 1, 65535
    timeout_input = input("\nТаймаут для каждого порта в секундах (по умолчанию 1): ").strip()
    timeout = float(timeout_input) if timeout_input else 1.0
    threads_input = input("Количество потоков (по умолчанию 500): ").strip()
    max_threads = int(threads_input) if threads_input else 500
    print("\nПолучение информации о хосте...")
    print("-" * 40)
    host_info = get_host_info(ip)
    print(f"IP: {ip}")
    print(f"Имя хоста: {host_info['Имя хоста']}")
    print(f"Статус: {host_info['Статус']}")
    print(f"\nБудет просканировано портов: {end_port - start_port + 1}")
    confirm = input("\nНачать сканирование? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Сканирование отменено.")
        return
    open_ports = scan_ports(ip, start_port, end_port, max_threads, timeout)
    if open_ports:
        print("\nСОХРАНЕНИЕ РЕЗУЛЬТАТОВ")
        print("-" * 40)
        filename = save_to_excel(ip, open_ports, host_info, (start_port, end_port))
        print(f"\nРезультаты сохранены в файл: {filename}")
        print("\nСВОДКА:")
        print("-" * 40)
        print(f"IP адрес: {ip}")
        print(f"Имя хоста: {host_info['Имя хоста']}")
        print(f"Диапазон портов: {start_port}-{end_port}")
        print(f"Всего портов: {end_port - start_port + 1}")
        print(f"Открытых портов: {len(open_ports)}")
        if len(open_ports) > 0:
            print("\nОткрытые порты:")
            for port, service in sorted(open_ports):
                print(f"  Порт {port}: {service}")
    else:
        print("\nОткрытых портов не обнаружено.")
if __name__ == "__main__":
    try:
        import pandas
        from openpyxl import Workbook
    except ImportError as e:
        print(f"Ошибка импорта: {e}")
        print("\nУстановите необходимые модули:")
        print("pip install pandas openpyxl")
        exit(1)
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nСканирование прервано пользователем.")
    except Exception as e:
        print(f"\nПроизошла ошибка: {e}")
        import traceback
        traceback.print_exc()