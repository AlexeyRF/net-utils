import ipaddress
import subprocess
import socket
import psutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import concurrent.futures
from datetime import datetime
import re
def sanitize_sheet_name(name):
    invalid_chars = r'[:\\/*?\[\]]'
    sanitized = re.sub(invalid_chars, '', name)
    return sanitized[:31] if sanitized else "Sheet"
def get_active_adapters():
    adapters = []
    for iface, addrs in psutil.net_if_addrs().items():
        stats = psutil.net_if_stats()
        if iface not in stats or not stats[iface].isup:
            continue
        ipv4_info = None
        for addr in addrs:
            if addr.family == socket.AF_INET and not addr.address.startswith('127.'):
                if not addr.address.startswith('169.254.'):
                    ipv4_info = {
                        'interface': iface,
                        'ip': addr.address,
                        'netmask': addr.netmask
                    }
                break
        if ipv4_info:
            adapters.append(ipv4_info)
    return adapters
def calculate_network_address(ip, netmask):
    try:
        network = ipaddress.IPv4Network(f"{ip}/{netmask}", strict=False)
        hosts = list(network.hosts())
        return {
            'network': str(network.network_address),
            'broadcast': str(network.broadcast_address),
            'total_hosts': network.num_addresses - 2,
            'all_hosts': [str(host) for host in hosts]
        }
    except Exception as e:
        print(f"Ошибка расчета сети для {ip}/{netmask}: {e}")
        return None
def ping_host(ip, timeout=1):
    try:
        if os.name == 'nt':
            result = subprocess.run(
                ['ping', '-n', '1', '-w', str(timeout * 1000), ip],
                capture_output=True,
                text=True,
                encoding='cp866',
                creationflags=subprocess.CREATE_NO_WINDOW
            )
            return ip, result.returncode == 0 and "TTL=" in result.stdout
        else:
            result = subprocess.run(
                ['ping', '-c', '1', '-W', str(timeout), ip],
                capture_output=True,
                text=True
            )
            return ip, result.returncode == 0
    except Exception as e:
        return ip, False
def parallel_ping_scan(ip_list, max_workers=50, timeout=1):
    responsive_hosts = []
    print(f"Начинаем параллельное сканирование ({max_workers} потоков)")
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_ip = {executor.submit(ping_host, ip, timeout): ip for ip in ip_list}
        completed = 0
        total = len(ip_list)
        for future in concurrent.futures.as_completed(future_to_ip):
            completed += 1
            ip, is_alive = future.result()
            progress = (completed / total) * 100
            print(f"Прогресс: {completed}/{total} ({progress:.1f}%) - {ip} {'активен' if is_alive else 'неактивен'}", end='\r')
            if is_alive:
                responsive_hosts.append(ip)
    print(f"\nСканирование завершено. Найдено активных хостов: {len(responsive_hosts)}")
    return responsive_hosts
def get_additional_info(ip):
    try:
        hostname = socket.gethostbyaddr(ip)[0] if ip != '127.0.0.1' else 'localhost'
    except:
        hostname = "Не определено"
    vendor_info = get_vendor_by_ip(ip)
    return {
        'hostname': hostname,
        'vendor': vendor_info
    }
def get_vendor_by_ip(ip):
    return "Не определено"
def scan_adapter_network(adapter, scan_large_networks=True):
    print(f"\n{'-'*60}")
    print(f"Обработка адаптера: {adapter['interface']}")
    print(f"IP: {adapter['ip']}, Маска: {adapter['netmask']}")
    network_info = calculate_network_address(adapter['ip'], adapter['netmask'])
    if not network_info:
        print(f"Ошибка при расчете сети для адаптера {adapter['interface']}")
        return None
    print(f"Сеть: {network_info['network']}")
    print(f"Доступные хосты: {network_info['total_hosts']}")
    if network_info['total_hosts'] > 256 and not scan_large_networks:
        answer = input(f"\nВ сети {adapter['interface']} более 256 хостов ({network_info['total_hosts']}).\n"
                      "Выполнить сканирование? (y/n): ")
        if answer.lower() != 'y':
            print("Пропускаем сканирование этой сети")
            return None
    if network_info['total_hosts'] == 0:
        print("Нет доступных хостов для сканирования")
        return None
    max_workers = min(100, max(10, network_info['total_hosts'] // 10))
    responsive_ips = parallel_ping_scan(network_info['all_hosts'], max_workers)
    return {
        'adapter_info': adapter,
        'network_info': network_info,
        'responsive_ips': responsive_ips
    }
def save_to_excel(results, filename="network_scan_results.xlsx"):
    if not results:
        print("Нет результатов для сохранения")
        return
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    results_summary = []
    for result in results:
        if not result or not result['responsive_ips']:
            continue
        adapter = result['adapter_info']
        network = result['network_info']
        responsive_ips = result['responsive_ips']
        detailed_results = []
        print(f"\nПолучение дополнительной информации для {len(responsive_ips)} хостов...")
        for ip in responsive_ips:
            info = get_additional_info(ip)
            detailed_results.append({
                'IP адрес': ip,
                'Имя хоста': info['hostname'],
                'Производитель': info['vendor'],
                'Время сканирования': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
        df = pd.DataFrame(detailed_results)
        sheet_name = sanitize_sheet_name(f"{adapter['interface']}")
        ws = wb.create_sheet(title=sheet_name)
        ws.append([f"Дата сканирования: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])
        ws.append(["Информация о сетевом адаптере:"])
        ws.append([f"Имя адаптера: {adapter['interface']}"])
        ws.append([f"IP адрес: {adapter['ip']}"])
        ws.append([f"Маска подсети: {adapter['netmask']}"])
        ws.append([f"Сеть: {network['network']}"])
        ws.append([f"Всего хостов в сети: {network['total_hosts']}"])
        ws.append([f"Найдено активных хостов: {len(responsive_ips)}"])
        ws.append([])
        headers = list(df.columns)
        ws.append(headers)
        for _, row in df.iterrows():
            ws.append([row[col] for col in headers])
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        ws.auto_filter.ref = f"A{5 + len(headers)}:D{5 + len(headers) + len(df)}"
        results_summary.append({
            'Адаптер': adapter['interface'],
            'IP': adapter['ip'],
            'Маска': adapter['netmask'],
            'Сеть': network['network'],
            'Всего хостов': network['total_hosts'],
            'Найдено': len(responsive_ips)
        })
    if results_summary:
        ws_summary = wb.create_sheet(title="Итог", index=0)
        ws_summary.append(["Сводный отчет"])
        ws_summary.append([f"Дата сканирования: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_summary.append([])
        summary_df = pd.DataFrame(results_summary)
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws_summary.append(r)
        for col_idx, col in enumerate(ws_summary.columns, 1):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column].width = adjusted_width
        ws_summary.auto_filter.ref = f"A4:F{4 + len(summary_df)}"
        ws_summary.append([])
        total_hosts = sum(item['Всего хостов'] for item in results_summary)
        total_found = sum(item['Найдено'] for item in results_summary)
        ws_summary.append(["ИТОГО:"])
        ws_summary.append([f"Всего адаптеров: {len(results_summary)}"])
        ws_summary.append([f"Всего хостов в сетях: {total_hosts}"])
        ws_summary.append([f"Найдено активных хостов: {total_found}"])
        ws_summary.append([f"Процент обнаружения: {(total_found/total_hosts*100 if total_hosts > 0 else 0):.2f}%"])
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"network_scan_{timestamp}.xlsx"
    wb.save(filename)
    print(f"\n{'-'*60}")
    print(f"Результаты сохранены в файл: {filename}")
    return filename
def main():
    print("=" * 60)
    print("           СКАНЕР СЕТИ v2.0")
    print("=" * 60)
    print()
    adapters = get_active_adapters()
    if not adapters:
        print("Не найдено активных сетевых адаптеров с IPv4 адресами")
        return
    print(f"Найдено активных адаптеров: {len(adapters)}")
    for i, adapter in enumerate(adapters, 1):
        print(f"  {i}. {adapter['interface']}: {adapter['ip']}/{adapter['netmask']}")
    print("\n" + "=" * 60)
    print("Настройки сканирования:")
    print("=" * 60)
    scan_large = input("Сканировать сети с более чем 256 хостов без подтверждения? (y/n): ").lower() == 'y'
    custom_timeout = input("Таймаут пинга в секундах (по умолчанию 1): ").strip()
    timeout = float(custom_timeout) if custom_timeout else 1.0
    print(f"\n{'-'*60}")
    print("Начинаем сканирование сетей...")
    print("-" * 60)
    all_results = []
    for adapter in adapters:
        result = scan_adapter_network(adapter, scan_large)
        if result:
            all_results.append(result)
    if all_results:
        filename = save_to_excel(all_results)
        print("\n" + "=" * 60)
        print("ИТОГОВАЯ СТАТИСТИКА:")
        print("=" * 60)
        total_hosts = sum(r['network_info']['total_hosts'] for r in all_results if r)
        total_found = sum(len(r['responsive_ips']) for r in all_results if r)
        print(f"Всего адаптеров: {len(all_results)}")
        print(f"Всего хостов в сетях: {total_hosts}")
        print(f"Найдено активных хостов: {total_found}")
        print(f"Процент обнаружения: {(total_found/total_hosts*100 if total_hosts > 0 else 0):.2f}%")
        print("\nДетали по адаптерам:")
        for result in all_results:
            if result:
                adapter = result['adapter_info']
                network = result['network_info']
                found = len(result['responsive_ips'])
                print(f"  {adapter['interface']}: {found}/{network['total_hosts']} хостов")
    else:
        print("\nНет данных для сохранения")
if __name__ == "__main__":
    try:
        import psutil
        import pandas
        from openpyxl import Workbook
    except ImportError as e:
        print(f"Ошибка импорта: {e}")
        print("\nУстановите необходимые модули:")
        print("pip install psutil pandas openpyxl")
        exit(1)
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nСканирование прервано пользователем")
    except Exception as e:
        print(f"\nПроизошла ошибка: {e}")
        import traceback
        traceback.print_exc()